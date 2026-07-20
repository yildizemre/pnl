from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

LOGO_PATHS = [
    Path(__file__).parent.parent / "src" / "public" / "siyah.png",
    Path(__file__).parent.parent / "public" / "siyah.png",
    Path(__file__).parent.parent / "src" / "public" / "beyaz.png",
]


def _logo_path():
    for p in LOGO_PATHS:
        if p.exists():
            return str(p)
    return None


def _draw_header(c, title: str, meta: dict | None, w, h):
    logo = _logo_path()
    y = h - 2 * cm
    if logo:
        try:
            c.drawImage(
                logo, 2 * cm, y - 1.2 * cm, width=4 * cm, height=1.2 * cm,
                preserveAspectRatio=True, mask="auto",
            )
            y -= 1.6 * cm
        except Exception:
            pass
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, y, "HypeVision AI Analytics")
    y -= 0.7 * cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, y, title[:80])
    y -= 0.55 * cm
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.grey)
    c.drawString(2 * cm, y, f"Olusturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    if meta:
        y -= 0.4 * cm
        line = " | ".join(
            str(meta.get(k) or "")
            for k in ("kurulum", "baslangic", "bitis", "period")
            if meta.get(k)
        )
        if line:
            c.drawString(2 * cm, y, line[:95])
    y -= 0.9 * cm
    c.setFillColor(colors.black)
    return y


def build_pdf(title: str, rows: list[dict], meta: dict | None = None) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    y = _draw_header(c, title, meta, w, h)
    c.setFont("Helvetica-Bold", 9)
    if rows:
        headers = list(rows[0].keys())[:5]
        for i, hdr in enumerate(headers):
            c.drawString(2 * cm + i * 3.5 * cm, y, str(hdr)[:18])
        y -= 0.45 * cm
        c.setFont("Helvetica", 8)
        for row in rows[:55]:
            if y < 2 * cm:
                c.showPage()
                y = h - 2 * cm
                c.setFont("Helvetica", 8)
            for i, key in enumerate(headers):
                c.drawString(2 * cm + i * 3.5 * cm, y, str(row.get(key, ""))[:22])
            y -= 0.4 * cm
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.grey)
    c.drawString(2 * cm, 1.5 * cm, "HypeVision — panel.hypevisionlab.com")
    c.save()
    buf.seek(0)
    return buf.read()


def build_isg_weekly_pdf(summary: dict, meta: dict | None = None) -> bytes:
    title = f"Haftalik ISG Ozeti ({summary.get('baslangic')} — {summary.get('bitis')})"
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    m = {**(meta or {}), "baslangic": summary.get("baslangic"), "bitis": summary.get("bitis")}
    y = _draw_header(c, title, m, w, h)

    c.setFont("Helvetica-Bold", 11)
    c.drawString(2 * cm, y, f"Toplam olay: {summary.get('toplam', 0)}   Kritik: {summary.get('kritik', 0)}")
    y -= 0.7 * cm

    aks = summary.get("aksiyonlar") or {}
    c.setFont("Helvetica", 9)
    c.drawString(
        2 * cm, y,
        f"Acik: {aks.get('acik', 0)} | Kapandi: {aks.get('kapandi', 0)} | "
        f"Yanlis alarm: {aks.get('yanlis_alarm', 0)} | Egitim: {aks.get('egitim', 0)}",
    )
    y -= 0.8 * cm

    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y, "Gunluk trend")
    y -= 0.45 * cm
    c.setFont("Helvetica", 8)
    for row in summary.get("gunluk") or []:
        if y < 2.5 * cm:
            c.showPage()
            y = h - 2 * cm
        c.drawString(2 * cm, y, f"{row.get('tarih')}: {row.get('adet')} olay")
        y -= 0.35 * cm

    y -= 0.4 * cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y, "Kategoriler")
    y -= 0.45 * cm
    c.setFont("Helvetica", 8)
    for row in (summary.get("kategoriler") or [])[:20]:
        if y < 2.5 * cm:
            c.showPage()
            y = h - 2 * cm
        c.drawString(2 * cm, y, f"{row.get('kategori')}: {row.get('adet')}")
        y -= 0.35 * cm

    y -= 0.4 * cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y, "Kameralar")
    y -= 0.45 * cm
    c.setFont("Helvetica", 8)
    for row in (summary.get("kameralar") or [])[:15]:
        if y < 2.5 * cm:
            c.showPage()
            y = h - 2 * cm
        c.drawString(2 * cm, y, f"{str(row.get('kamera', ''))[:40]}: {row.get('adet')}")
        y -= 0.35 * cm

    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.grey)
    c.drawString(2 * cm, 1.5 * cm, "HypeVision — Haftalik ISG PDF")
    c.save()
    buf.seek(0)
    return buf.read()


def build_xlsx(title: str, rows: list[dict], sheet_name: str = "Rapor") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(["HypeVision AI Analytics"])
    ws.append([title])
    ws.append([f"Olusturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
    ws.append([])
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_workbook_multi(sheets: list[tuple[str, list[dict]]], title: str = "HypeVision") -> bytes:
    wb = Workbook()
    first = True
    for name, rows in sheets:
        if first:
            ws = wb.active
            ws.title = name[:31]
            first = False
        else:
            ws = wb.create_sheet(name[:31])
        ws.append([title])
        ws.append([f"Olusturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
        ws.append([])
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
